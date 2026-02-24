"""
════════════════════════════════════════════════════════════════════════════════
  LIBRO DE CAJA PRO PYME — Conforme SII Chile (Art. 14 D N°3 y N°8)
  Desarrollado para cumplir normativa vigente SII
════════════════════════════════════════════════════════════════════════════════
  Instalación:
      pip install streamlit pandas openpyxl

  Ejecución:
      streamlit run libro_caja_app.py
════════════════════════════════════════════════════════════════════════════════
"""

import io
import re
import os
from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter
from fpdf import FPDF


# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTES
# ─────────────────────────────────────────────────────────────────────────────
TIPO_DOC_NOMBRES = {
    33: "Factura Electrónica",
    34: "Factura No Afecta o Exenta Elec.",
    35: "Boleta Afecta Electrónica",
    38: "Boleta No Afecta o Exenta Elec.",
    39: "Boleta Electrónica",
    41: "Boleta Exenta Electrónica",
    46: "Factura de Compra Electrónica",
    48: "Comprobante de Pago Electrónico",
    52: "Guía de Despacho Electrónica",
    56: "Nota de Débito Electrónica",
    61: "Nota de Crédito Electrónica",
    110: "Factura de Exportación Electrónica",
    111: "Nota de Débito de Exportación Elec.",
    112: "Nota de Crédito de Exportación Elec.",
}

BOLETAS_AFECTAS = [35, 39]
BOLETAS_EXENTAS = [38, 41]
COMPROBANTES_PAGO = [48]
NOTAS_CREDITO = [61, 112]
NOTAS_DEBITO = [56, 111]
FACTURAS_VENTA = [33, 34, 110]

SEPARADORES = [";", ",", "\t", "|"]
ENCODINGS = ["utf-8", "utf-8-sig", "latin-1", "iso-8859-1", "cp1252"]


# ─────────────────────────────────────────────────────────────────────────────
# UTILIDADES DE LECTURA CSV
# ─────────────────────────────────────────────────────────────────────────────
def detectar_separador(contenido: bytes) -> str:
    muestra = contenido[:2000].decode("latin-1", errors="ignore")
    conteos = {sep: muestra.count(sep) for sep in SEPARADORES}
    return max(conteos, key=conteos.get)


def leer_csv(archivo) -> pd.DataFrame:
    """
    Lee un CSV del SII detectando separador y encoding.
    Maneja trailing semicolons que generan los exportadores del SII.
    """
    contenido = archivo.read()
    sep = detectar_separador(contenido)

    for enc in ENCODINGS:
        try:
            texto = contenido.decode(enc)
            lines = texto.splitlines()
            if not lines:
                continue

            n_header = len(lines[0].split(sep))
            n_data = len(lines[1].split(sep)) if len(lines) > 1 else n_header

            if n_data > n_header:
                extras = n_data - n_header
                header_ext = lines[0].strip().split(sep) + [f"_extra_{i}" for i in range(extras)]
                df = pd.read_csv(
                    io.BytesIO(contenido),
                    sep=sep,
                    encoding=enc,
                    dtype=str,
                    names=header_ext,
                    skiprows=1,
                    on_bad_lines="skip",
                )
            else:
                df = pd.read_csv(
                    io.BytesIO(contenido),
                    sep=sep,
                    encoding=enc,
                    dtype=str,
                    on_bad_lines="skip",
                )

            df.columns = df.columns.str.strip()
            df = df[[c for c in df.columns if not str(c).startswith("_extra_")]]
            return df
        except Exception:
            continue

    raise ValueError(f"No se pudo leer el archivo: {getattr(archivo, 'name', 'desconocido')}")


def parsear_fecha(valor: str) -> pd.Timestamp | None:
    if not valor or pd.isna(valor):
        return None
    valor = str(valor).strip()
    # Solo tomar la parte de fecha (dd/mm/aaaa) si viene con hora
    parte_fecha = valor.split(" ")[0]
    for fmt in ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"]:
        try:
            return pd.to_datetime(parte_fecha, format=fmt)
        except Exception:
            pass
    return None


def a_numero(valor) -> float:
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return 0.0
    try:
        return float(str(valor).replace(".", "").replace(",", ".").strip())
    except Exception:
        return 0.0


# ─────────────────────────────────────────────────────────────────────────────
# PROCESAMIENTO VENTAS
# ─────────────────────────────────────────────────────────────────────────────
def procesamiento_ventas(archivos_ventas: list, archivos_resumen: list, periodo: str = "") -> pd.DataFrame:
    """
    Procesa archivos CSV de ventas (facturas y resúmenes).
    Retorna DataFrame con columnas del Libro de Caja.
    """
    registros = []

    # ── Facturas de venta ────────────────────────────────────────────────────
    for archivo in archivos_ventas:
        df = leer_csv(archivo)
        col_map = _mapear_columnas_ventas(df)
        if col_map is None:
            st.warning(f"⚠️ No se reconocieron columnas en {archivo.name}. Se omite.")
            continue

        # Pre-scan: construir diccionario folio → fecha para filas duplicadas sin fecha
        fechas_por_folio = {}
        for _, f in df.iterrows():
            folio_tmp = str(f.get(col_map.get("folio", ""), "")).strip()
            fecha_tmp = parsear_fecha(f.get(col_map.get("fecha", ""), ""))
            if folio_tmp and fecha_tmp is not None:
                fechas_por_folio[folio_tmp] = fecha_tmp

        for _, fila in df.iterrows():
            tipo_doc_raw = a_numero(fila.get(col_map.get("tipo_doc", ""), 0))
            tipo_doc = int(tipo_doc_raw) if tipo_doc_raw else 0

            # Solo procesar facturas en este archivo
            if tipo_doc not in FACTURAS_VENTA + NOTAS_CREDITO + NOTAS_DEBITO:
                continue

            folio = str(fila.get(col_map.get("folio", ""), "")).strip()

            fecha = parsear_fecha(fila.get(col_map.get("fecha", ""), ""))
            if fecha is None:
                # Buscar fecha del mismo documento (folio)
                fecha = fechas_por_folio.get(folio)
            if fecha is None:
                continue

            rut_cliente = str(fila.get(col_map.get("rut", ""), "")).strip()
            razon = str(fila.get(col_map.get("razon", ""), "")).strip()
            monto_neto = a_numero(fila.get(col_map.get("neto", ""), 0))
            monto_exento = a_numero(fila.get(col_map.get("exento", ""), 0))
            monto_total = a_numero(fila.get(col_map.get("total", ""), 0))


            # Determinar tipo operación y montos según tipo de documento
            if tipo_doc in NOTAS_CREDITO:
                tipo_op = 2  # Egreso (devuelve dinero)
                c8 = abs(monto_total)
                c9 = abs(monto_neto + monto_exento)
                glosa = f"NC Venta — {razon}"
            elif tipo_doc in NOTAS_DEBITO:
                tipo_op = 1  # Ingreso adicional
                c8 = abs(monto_total)
                c9 = abs(monto_neto + monto_exento)
                glosa = f"ND Venta — {razon}"
            else:
                tipo_op = 1
                c8 = abs(monto_total)
                c9 = abs(monto_neto + monto_exento)  # Sin IVA en ProPyme
                glosa = f"Venta — {razon}"

            registros.append({
                "Tipo Operación": tipo_op,
                "N° Documento": folio,
                "Tipo Documento": tipo_doc,
                "RUT Emisor": rut_cliente,
                "Fecha Operación": fecha,
                "Glosa de Operación": glosa,
                "C8": c8,
                "C9": c9,
                "_origen": "venta_factura",
            })

    # ── Resúmenes de ventas (boletas / comprobantes) ──────────────────────────
    for archivo in archivos_resumen:
        df = leer_csv(archivo)
        _procesar_resumen_ventas(df, archivo.name, registros, periodo)

    if not registros:
        return pd.DataFrame()

    return pd.DataFrame(registros)


def _mapear_columnas_ventas(df: pd.DataFrame) -> dict | None:
    """Mapea columnas del CSV de ventas a nombres estándar."""
    cols = {c.lower(): c for c in df.columns}
    mapa = {}

    # Tipo documento
    for k in ["tipo doc", "tipo_doc", "tipodoc", "tipo documento"]:
        if k in cols:
            mapa["tipo_doc"] = cols[k]
            break

    # Folio
    for k in ["folio", "n° folio", "numero folio"]:
        if k in cols:
            mapa["folio"] = cols[k]
            break

    # Fecha
    for k in ["fecha docto", "fecha_docto", "fechadocto", "fecha operación", "fecha"]:
        if k in cols:
            mapa["fecha"] = cols[k]
            break

    # RUT
    for k in ["rut cliente", "rut_cliente", "rutcliente", "rut proveedor", "rut"]:
        if k in cols:
            mapa["rut"] = cols[k]
            break

    # Razón social
    for k in ["razon social", "razón social", "razon_social"]:
        if k in cols:
            mapa["razon"] = cols[k]
            break

    # Montos
    for k in ["monto neto", "monto_neto", "neto"]:
        if k in cols:
            mapa["neto"] = cols[k]
            break

    for k in ["monto exento", "monto_exento", "exento"]:
        if k in cols:
            mapa["exento"] = cols[k]
            break

    for k in ["monto total", "monto_total", "total"]:
        if k in cols:
            mapa["total"] = cols[k]
            break

    if "tipo_doc" not in mapa or "fecha" not in mapa:
        return None
    return mapa


def _fecha_fallback_desde_nombre(nombre_archivo: str, periodo: str) -> pd.Timestamp:
    """
    Intenta extraer año y mes del nombre del archivo.
    Busca patrones como: 2024-11, 202411, nov2024, etc.
    Prioriza los números al final del archivo.
    Si no encuentra nada, usa el 31/12 del año del período.
    """
    # Buscar patrón YYYY-MM o YYYYMM en el nombre del archivo
    # Usamos findall y leemos al revés para evitar confundir fechas con el RUT
    matches = re.findall(r"(20[0-9]{2})[_\-]?([0-9]{2})", nombre_archivo)
    if matches:
        for m in reversed(matches):
            anio = int(m[0])
            mes = int(m[1])
            if 2000 <= anio <= 2100 and 1 <= mes <= 12:
                # Último día del mes
                import calendar
                ultimo_dia = calendar.monthrange(anio, mes)[1]
                return pd.Timestamp(f"{anio}-{mes:02d}-{ultimo_dia}")

    # Fallback: último día del año del período ingresado por el usuario
    anio_periodo = str(periodo).strip()
    if anio_periodo.isdigit() and len(anio_periodo) == 4:
        return pd.Timestamp(f"{anio_periodo}-12-31")

    # Último recurso: año actual
    return pd.Timestamp(f"{datetime.now().year}-12-31")


def _procesar_resumen_ventas(df: pd.DataFrame, nombre: str, registros: list, periodo: str = ""):
    """
    Procesa archivo resumen de ventas (boletas y comprobantes por tipo y día).
    Formato esperado: Tipo Documento | Total Documentos | Monto Exento | Monto Neto | Monto IVA | Monto Total
    Si el CSV tiene fecha (detalle diario), agrupa por fecha.
    Si es resumen mensual, extrae una fila por tipo de documento.
    """
    cols = {c.lower().strip(): c for c in df.columns}

    # Detectar si hay columna de fecha (detalle diario)
    col_fecha = None
    for k in ["fecha", "fecha docto", "fecha_docto"]:
        if k in cols:
            col_fecha = cols[k]
            break

    col_tipo = None
    for k in ["tipo documento", "tipo_documento", "tipodocumento"]:
        if k in cols:
            col_tipo = cols[k]
            break

    col_neto = None
    for k in ["monto neto", "monto_neto", "neto"]:
        if k in cols:
            col_neto = cols[k]
            break

    col_exento = None
    for k in ["monto exento", "monto_exento", "exento"]:
        if k in cols:
            col_exento = cols[k]
            break

    col_total = None
    for k in ["monto total", "monto_total", "total"]:
        if k in cols:
            col_total = cols[k]
            break

    # Detectar folio inicio / fin (para resúmenes diarios con rango)
    col_folio_ini = None
    col_folio_fin = None
    for k in ["folio inicial", "folio_inicial", "desde"]:
        if k in cols:
            col_folio_ini = cols[k]
            break
    for k in ["folio final", "folio_final", "hasta"]:
        if k in cols:
            col_folio_fin = cols[k]
            break

    for _, fila in df.iterrows():
        tipo_str = str(fila.get(col_tipo, "") if col_tipo else "").strip()
        if not tipo_str:
            continue

        # Extraer código numérico del tipo documento ej: "Boleta Afecta Electrónica(35)"
        match = re.search(r"\((\d+)\)", tipo_str)
        if not match:
            # Intentar leer directamente si es número
            try:
                codigo = int(tipo_str)
            except Exception:
                continue
        else:
            codigo = int(match.group(1))

        # Ignorar facturas en resúmenes (se procesan aparte)
        if codigo in FACTURAS_VENTA:
            continue

        # Solo boletas y comprobantes de pago
        # Solo boletas y comprobantes de pago (Excluir NC 61 y ND)
        if codigo not in BOLETAS_AFECTAS + BOLETAS_EXENTAS + COMPROBANTES_PAGO:
            continue

        monto_neto = a_numero(fila.get(col_neto, 0) if col_neto else 0)
        monto_exento = a_numero(fila.get(col_exento, 0) if col_exento else 0)
        monto_total = a_numero(fila.get(col_total, 0) if col_total else 0)


        # Fecha
        # Solicitado explícitamente: trasladar el periodo por su último día del mes
        # cortando el nombre del archivo, solo en la subida de las boletas.
        fecha = _fecha_fallback_desde_nombre(nombre, periodo)

        if fecha is None:
            continue

        # N° Documento
        if col_folio_ini and col_folio_fin:
            f_ini = str(fila.get(col_folio_ini, "")).strip()
            f_fin = str(fila.get(col_folio_fin, "")).strip()
            n_doc = f"{f_ini} al {f_fin}" if f_ini and f_fin else "Z"
        else:
            n_doc = "Z"

        # Como excluimos NC y ND, todo lo restante es venta/ingreso
        tipo_op = 1
        c9 = abs(monto_neto + monto_exento)

        registros.append({
            "Tipo Operación": tipo_op,
            "N° Documento": n_doc,
            "Tipo Documento": codigo,
            "RUT Emisor": "",
            "Fecha Operación": fecha,
            "Glosa de Operación": f"Resumen ventas boletas del día — {TIPO_DOC_NOMBRES.get(codigo, tipo_str)}",
            "C8": abs(monto_total),
            "C9": c9,
            "_origen": "resumen_ventas",
        })


# ─────────────────────────────────────────────────────────────────────────────
# PROCESAMIENTO DATOS F29 (PEGADO DE TEXTO)
# ─────────────────────────────────────────────────────────────────────────────

def procesar_texto_f29(texto: str) -> pd.DataFrame:
    """
    Procesa datos de pagos F29 pegados como texto.
    Formato esperado (separado por comas):
        C2, C3 (Folio), C4 (Descripción), C6 (Fecha), C7 (Detalle), C9 (Monto)
    Ejemplo:
        2,8091536376,Formulario F-29,06/02/2025,Pago del F-29,"$151,077"
    Retorna un DataFrame con formato de Libro de Caja.
    """
    if not texto or not texto.strip():
        return pd.DataFrame()

    data = []
    lineas = texto.strip().splitlines()

    # Detectar si la primera línea es un encabezado
    primera = lineas[0].strip().lower()
    if "folio" in primera or "descripción" in primera or "descripcion" in primera or "monto" in primera or "c2" in primera:
        lineas = lineas[1:]  # saltar encabezado

    for num_linea, linea in enumerate(lineas, start=1):
        linea = linea.strip()
        if not linea:
            continue

        # Parsear con csv.reader para manejar comillas y comas dentro de valores
        import csv
        try:
            partes = list(csv.reader([linea]))[0]
        except Exception:
            st.warning(f"⚠️ Línea {num_linea}: no se pudo parsear → '{linea}'")
            continue

        # Limpiar espacios
        partes = [p.strip() for p in partes]

        if len(partes) < 6:
            st.warning(f"⚠️ Línea {num_linea}: se esperan 6 campos, se encontraron {len(partes)} → '{linea}'")
            continue

        try:
            tipo_op = int(partes[0])          # C2 — Tipo Operación (1=Ingreso, 2=Egreso)
            folio = partes[1].strip()          # C3 — N° Documento / Folio
            descripcion = partes[2].strip()    # C4 — Tipo Documento (ej: Formulario F-29)
            fecha_str = partes[3].strip()      # C6 — Fecha Presentación
            detalle = partes[4].strip()        # C7 — Glosa / Detalle
            monto_str = partes[5].strip()      # C9 — Monto

            # Parsear fecha
            fecha_dt = parsear_fecha(fecha_str)
            if fecha_dt is None:
                st.warning(f"⚠️ Línea {num_linea}: fecha inválida '{fecha_str}'")
                continue

            # Limpiar monto: quitar $, puntos de miles, espacios
            monto_limpio = monto_str.replace("$", "").replace(".", "").replace(",", "").replace(" ", "").strip()
            if not monto_limpio.lstrip("-").isdigit():
                st.warning(f"⚠️ Línea {num_linea}: monto inválido '{monto_str}'")
                continue
            monto = abs(int(monto_limpio))

            data.append({
                "Tipo Operación": tipo_op,
                "N° Documento": folio,
                "Tipo Documento": descripcion,
                "RUT Emisor": "",
                "Fecha Operación": fecha_dt,
                "Glosa de Operación": detalle,
                "C8": monto,
                "C9": 0,
                "_origen": "f29_texto"
            })

        except ValueError as e:
            st.warning(f"⚠️ Línea {num_linea}: error de valor → {e}")
            continue

    if not data:
        st.warning("⚠️ No se encontraron datos válidos en el texto pegado.")
        return pd.DataFrame()

    st.success(f"✅ Se procesaron {len(data)} registros de F29 desde el texto.")
    return pd.DataFrame(data)


# ─────────────────────────────────────────────────────────────────────────────
# PROCESAMIENTO COMPRAS
# ─────────────────────────────────────────────────────────────────────────────
def procesamiento_compras(archivos_compras: list) -> pd.DataFrame:
    """Procesa archivos CSV de compras. Retorna DataFrame Libro de Caja."""
    registros = []

    for archivo in archivos_compras:
        df = leer_csv(archivo)
        col_map = _mapear_columnas_compras(df)
        if col_map is None:
            st.warning(f"⚠️ No se reconocieron columnas en {archivo.name}. Se omite.")
            continue

        # Pre-scan: construir diccionario folio → fecha para filas duplicadas sin fecha
        fechas_por_folio = {}
        for _, f in df.iterrows():
            folio_tmp = str(f.get(col_map.get("folio", ""), "")).strip()
            fecha_tmp = parsear_fecha(f.get(col_map.get("fecha", ""), ""))
            if folio_tmp and fecha_tmp is not None:
                fechas_por_folio[folio_tmp] = fecha_tmp

        for _, fila in df.iterrows():
            tipo_doc_raw = a_numero(fila.get(col_map.get("tipo_doc", ""), 0))
            tipo_doc = int(tipo_doc_raw) if tipo_doc_raw else 0

            folio = str(fila.get(col_map.get("folio", ""), "")).strip()

            fecha = parsear_fecha(fila.get(col_map.get("fecha", ""), ""))
            if fecha is None:
                # Buscar fecha del mismo documento (folio)
                fecha = fechas_por_folio.get(folio)
            if fecha is None:
                continue
            rut_prov = str(fila.get(col_map.get("rut", ""), "")).strip()
            razon = str(fila.get(col_map.get("razon", ""), "")).strip()
            monto_neto = a_numero(fila.get(col_map.get("neto", ""), 0))
            monto_exento = a_numero(fila.get(col_map.get("exento", ""), 0))
            monto_total = a_numero(fila.get(col_map.get("total", ""), 0))

            # Columnas adicionales que afectan la base imponible en COMPRAS
            neto_activo_fijo = a_numero(fila.get(col_map.get("neto_activo_fijo", ""), 0))
            iva_no_recuperable = a_numero(fila.get(col_map.get("iva_no_recuperable", ""), 0))
            tab_puros = a_numero(fila.get(col_map.get("tab_puros", ""), 0))
            tab_cigarrillos = a_numero(fila.get(col_map.get("tab_cigarrillos", ""), 0))
            tab_elaborados = a_numero(fila.get(col_map.get("tab_elaborados", ""), 0))
            impto_sin_credito = a_numero(fila.get(col_map.get("impto_sin_credito", ""), 0))
            otro_impuesto = a_numero(fila.get(col_map.get("otro_impuesto", ""), 0))


            # C9 para compras: Neto + Exento + Neto Activo Fijo
            #   + IVA No Recuperable + Tabacos (puros+cigarrillos+elaborados)
            #   + Impto. Sin Derecho a Crédito + Valor Otro Impuesto
            base_imponible = (monto_neto + monto_exento
                              + neto_activo_fijo
                              + iva_no_recuperable
                              + tab_puros + tab_cigarrillos + tab_elaborados
                              + impto_sin_credito
                              + otro_impuesto)

            if tipo_doc in NOTAS_CREDITO:
                # NC de proveedor: reembolso → ingreso, aumenta base imponible
                tipo_op = 1
                c8 = abs(monto_total)
                c9 = abs(base_imponible)
                glosa = f"NC Compra — {razon}"
            elif tipo_doc in NOTAS_DEBITO:
                # ND de proveedor: pago adicional → egreso, disminuye base imponible
                tipo_op = 2
                c8 = abs(monto_total)
                c9 = abs(base_imponible)
                glosa = f"ND Compra — {razon}"
            else:
                tipo_op = 2  # Egreso
                c8 = abs(monto_total)
                c9 = abs(base_imponible)
                glosa = f"Compra — {razon}"

            registros.append({
                "Tipo Operación": tipo_op,
                "N° Documento": folio,
                "Tipo Documento": tipo_doc,
                "RUT Emisor": rut_prov,
                "Fecha Operación": fecha,
                "Glosa de Operación": glosa,
                "C8": c8,
                "C9": c9,
                "_origen": "compra",
            })

    if not registros:
        return pd.DataFrame()
    return pd.DataFrame(registros)


def _mapear_columnas_compras(df: pd.DataFrame) -> dict | None:
    cols = {c.lower().strip(): c for c in df.columns}
    mapa = {}

    for k in ["tipo doc", "tipo_doc", "tipodoc"]:
        if k in cols:
            mapa["tipo_doc"] = cols[k]
            break

    for k in ["folio", "n° folio"]:
        if k in cols:
            mapa["folio"] = cols[k]
            break

    for k in ["fecha docto", "fecha_docto", "fechadocto", "fecha recepcion"]:
        if k in cols:
            mapa["fecha"] = cols[k]
            break
    if "fecha" not in mapa:
        for k in ["fecha"]:
            if k in cols:
                mapa["fecha"] = cols[k]
                break

    for k in ["rut proveedor", "rut_proveedor", "rutproveedor"]:
        if k in cols:
            mapa["rut"] = cols[k]
            break

    for k in ["razon social", "razón social", "razon_social"]:
        if k in cols:
            mapa["razon"] = cols[k]
            break

    for k in ["monto neto", "monto_neto", "neto"]:
        if k in cols:
            mapa["neto"] = cols[k]
            break

    for k in ["monto exento", "monto_exento", "exento"]:
        if k in cols:
            mapa["exento"] = cols[k]
            break

    for k in ["monto total", "monto_total", "total"]:
        if k in cols:
            mapa["total"] = cols[k]
            break

    # ── Columnas adicionales que afectan Base Imponible en Compras ────────
    for k in ["monto neto activo fijo", "monto_neto_activo_fijo",
              "neto activo fijo", "neto_activo_fijo"]:
        if k in cols:
            mapa["neto_activo_fijo"] = cols[k]
            break

    for k in ["monto iva no recuperable", "monto_iva_no_recuperable",
              "iva no recuperable", "iva_no_recuperable",
              "monto iva no rec.", "iva no rec."]:
        if k in cols:
            mapa["iva_no_recuperable"] = cols[k]
            break

    for k in ["tabacos puros", "tabacos_puros",
              "monto imp. tab. puros", "tab. puros",
              "mto imp. tab. puros", "imp tab puros",
              "tabacos puros no nominados"]:
        if k in cols:
            mapa["tab_puros"] = cols[k]
            break

    for k in ["mto imp. tab. cigarrillos", "tab. cigarrillos",
              "tabacos cigarrillos", "cigarrillos",
              "imp tab cigarrillos"]:
        if k in cols:
            mapa["tab_cigarrillos"] = cols[k]
            break

    for k in ["mto imp. tab. elaborados", "tab. elaborados",
              "tabacos elaborados", "elaborados",
              "imp tab elaborados"]:
        if k in cols:
            mapa["tab_elaborados"] = cols[k]
            break

    for k in ["impto. sin derecho a crédito", "impto sin derecho a credito",
              "impto_sin_credito", "monto sin derecho a credito",
              "imp. sin derecho a crédito", "imp sin derecho a credito",
              "sin derecho a credito"]:
        if k in cols:
            mapa["impto_sin_credito"] = cols[k]
            break

    for k in ["valor otro impuesto", "otro impuesto", "otros impuestos",
              "otro_impuesto", "otros impuestos sin credito"]:
        if k in cols:
            mapa["otro_impuesto"] = cols[k]
            break

    if "tipo_doc" not in mapa or "fecha" not in mapa:
        return None
    return mapa


# ─────────────────────────────────────────────────────────────────────────────
# PROCESAMIENTO HONORARIOS (Excel)
# ─────────────────────────────────────────────────────────────────────────────

def procesar_texto_honorarios(texto: str) -> pd.DataFrame:
    """
    Procesa texto pegado con datos de boletas de honorarios.
    Formato esperado (CSV): C2,C3,C4,C5,C6,C7,C8,C9
    Ejemplo: 2,1389,Boleta de Honorarios Electrónica,13064992-0,05/09/2024,ALEJANDRO PATRICIO DUQUE SOTO,108684,108684
    """
    data = []
    lineas = texto.strip().splitlines()

    for num_linea, linea in enumerate(lineas, start=1):
        linea = linea.strip()
        if not linea:
            continue

        # Detectar y omitir líneas de encabezado
        linea_lower = linea.lower()
        if linea_lower.startswith("c2") or "tipo" in linea_lower and "operaci" in linea_lower:
            continue

        import csv
        try:
            partes = list(csv.reader([linea]))[0]
        except Exception:
            st.warning(f"⚠️ Línea {num_linea}: no se pudo parsear → '{linea}'")
            continue

        partes = [p.strip() for p in partes]

        if len(partes) < 8:
            st.warning(f"⚠️ Línea {num_linea}: se esperan 8 campos (C2-C9), se encontraron {len(partes)} → '{linea}'")
            continue

        try:
            tipo_op = int(partes[0])             # C2 — Tipo Operación (siempre 2 = Egreso)
            numero = partes[1].strip()            # C3 — N° Documento
            tipo_doc = partes[2].strip()          # C4 — Tipo Documento
            rut = partes[3].strip()               # C5 — RUT Emisor
            fecha_str = partes[4].strip()         # C6 — Fecha
            nombre = partes[5].strip()            # C7 — Nombre o Razón Social
            monto_c8_str = partes[6].strip()      # C8 — Pagado
            monto_c9_str = partes[7].strip()      # C9 — Bruto

            # Parsear fecha
            fecha_dt = parsear_fecha(fecha_str)
            if fecha_dt is None:
                st.warning(f"⚠️ Línea {num_linea}: fecha inválida '{fecha_str}'")
                continue

            # Limpiar montos
            c8_limpio = monto_c8_str.replace("$", "").replace(".", "").replace(",", "").replace(" ", "").strip()
            c9_limpio = monto_c9_str.replace("$", "").replace(".", "").replace(",", "").replace(" ", "").strip()

            c8 = abs(int(c8_limpio)) if c8_limpio.lstrip("-").isdigit() else 0
            c9 = abs(int(c9_limpio)) if c9_limpio.lstrip("-").isdigit() else 0

            if c8 == 0 and c9 == 0:
                continue

            data.append({
                "Tipo Operación": tipo_op,
                "N° Documento": numero,
                "Tipo Documento": tipo_doc if tipo_doc else "Boleta de Honorarios Electrónica",
                "RUT Emisor": rut,
                "Fecha Operación": fecha_dt,
                "Glosa de Operación": nombre,
                "C8": c8,
                "C9": c9,
                "_origen": "honorarios_texto"
            })

        except ValueError as e:
            st.warning(f"⚠️ Línea {num_linea}: error de valor → {e}")
            continue

    if not data:
        st.warning("⚠️ No se encontraron datos válidos de honorarios en el texto pegado.")
        return pd.DataFrame()

    st.success(f"✅ Se procesaron {len(data)} boletas de honorarios desde el texto.")
    return pd.DataFrame(data)


# ─────────────────────────────────────────────────────────────────────────────
# GENERACIÓN LIBRO DE CAJA
# ─────────────────────────────────────────────────────────────────────────────
def generar_libro_caja(
    df_ventas: pd.DataFrame,
    df_compras: pd.DataFrame,
    saldo_inicial: float,
    rut_empresa: str,
    nombre_empresa: str,
    periodo: str,
) -> pd.DataFrame:
    """Combina ventas y compras, ordena cronológicamente y agrega correlativo."""
    frames = [f for f in [df_ventas, df_compras] if not f.empty]

    # Fila saldo inicial
    saldo_row = pd.DataFrame([{
        "Tipo Operación": 0,
        "N° Documento": "",
        "Tipo Documento": "",
        "RUT Emisor": rut_empresa,
        "Tipo Documento": "",
        "RUT Emisor": rut_empresa,
        "Fecha Operación": pd.Timestamp(f"{str(periodo).strip()}-01-01") if periodo and str(periodo).strip().isdigit() and len(str(periodo).strip()) == 4 else pd.Timestamp.now().replace(month=1, day=1),
        "Glosa de Operación": "Saldo Inicial",
        "C8": float(saldo_inicial),
        "C9": 0.0,
        "_origen": "saldo_inicial",
    }])

    if frames:
        df_all = pd.concat(frames, ignore_index=True)
        df_all = df_all.sort_values("Fecha Operación", na_position="first")
        df_final = pd.concat([saldo_row, df_all], ignore_index=True)
    else:
        df_final = saldo_row

    df_final = df_final.reset_index(drop=True)
    df_final.insert(0, "N° Correlativo", range(1, len(df_final) + 1))

    return df_final


def calcular_totales(df: pd.DataFrame) -> dict:
    mask_ing = df["Tipo Operación"].isin([0, 1])
    mask_egr = df["Tipo Operación"] == 2

    total_ingresos = df.loc[mask_ing, "C8"].sum()
    total_egresos = df.loc[mask_egr, "C8"].sum()
    saldo_flujo = total_ingresos - total_egresos

    ing_bi = df.loc[mask_ing, "C9"].sum()
    egr_bi = df.loc[mask_egr, "C9"].sum()
    resultado_neto = ing_bi - egr_bi

    return {
        "total_ingresos": total_ingresos,
        "total_egresos": total_egresos,
        "saldo_flujo": saldo_flujo,
        "ing_bi": ing_bi,
        "egr_bi": egr_bi,
        "resultado_neto": resultado_neto,
    }


# ─────────────────────────────────────────────────────────────────────────────
# EXPORTACIÓN EXCEL (formato oficial SII)
# ─────────────────────────────────────────────────────────────────────────────
def exportar_excel(
    df: pd.DataFrame,
    totales: dict,
    rut_empresa: str,
    nombre_empresa: str,
    periodo: str,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Libro Caja"

    # ── Colores ────────────────────────────────────────────────────────────
    AZUL_OSCURO = "1F3864"
    AZUL_HEADER = "2E5090"
    AZUL_CLARO = "BDD7EE"
    GRIS_CLARO = "F2F2F2"
    AMARILLO = "FFFF00"
    VERDE = "E2EFDA"
    ROJO = "FCE4D6"
    BLANCO = "FFFFFF"

    borde_fino = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def estilo_celda(ws, fila, col, valor, negrita=False, alineacion="left",
                     fondo=None, fuente_color="000000", borde=True, num_fmt=None):
        cell = ws.cell(row=fila, column=col, value=valor)
        cell.font = Font(name="Arial", bold=negrita, color=fuente_color, size=9)
        cell.alignment = Alignment(horizontal=alineacion, vertical="center", wrap_text=True)
        if fondo:
            cell.fill = PatternFill("solid", start_color=fondo)
        if borde:
            cell.border = borde_fino
        if num_fmt:
            cell.number_format = num_fmt
        return cell

    # ── Título principal ───────────────────────────────────────────────────
    ws.merge_cells("A1:M1")
    titulo = ws["A1"]
    titulo.value = (
        "ANEXO 3. LIBRO DE CAJA CONTRIBUYENTES ACOGIDOS AL RÉGIMEN DEL "
        "ARTÍCULO 14 LETRA D) DEL N°3 Y N°8 LETRA (a) DE LA LEY SOBRE IMPUESTO A LA RENTA"
    )
    titulo.font = Font(name="Arial", bold=True, color=BLANCO, size=10)
    titulo.fill = PatternFill("solid", start_color=AZUL_OSCURO)
    titulo.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # ── Datos empresa ──────────────────────────────────────────────────────
    ws.merge_cells("A2:B2"); ws["A2"].value = "PERÍODO"
    ws.merge_cells("C2:F2"); ws["C2"].value = periodo
    ws.merge_cells("A3:B3"); ws["A3"].value = "RUT"
    ws.merge_cells("C3:F3"); ws["C3"].value = rut_empresa
    ws.merge_cells("A4:B4"); ws["A4"].value = "NOMBRE / RAZÓN SOCIAL"
    ws.merge_cells("C4:M4"); ws["C4"].value = nombre_empresa

    for row in range(2, 5):
        for col in range(1, 14):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(name="Arial", bold=(col <= 2), size=9)
            cell.fill = PatternFill("solid", start_color=GRIS_CLARO)
            cell.border = borde_fino
            cell.alignment = Alignment(vertical="center")

    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 18

    # ── Encabezados de sección ─────────────────────────────────────────────
    ws.merge_cells("A5:M5")
    ws["A5"].value = "REGISTRO DE OPERACIONES"
    ws["A5"].font = Font(name="Arial", bold=True, color=BLANCO, size=10)
    ws["A5"].fill = PatternFill("solid", start_color=AZUL_HEADER)
    ws["A5"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[5].height = 20

    # ── Encabezados de columna ─────────────────────────────────────────────
    headers = [
        ("A6", "N° CORRELATIVO\n(C1)"),
        ("B6", "TIPO OPERACIÓN\n(C2)"),
        ("C6", "N° DE DOCUMENTO\n(C3)"),
        ("D6", "TIPO DOCUMENTO\n(C4)"),
        ("E6", "RUT EMISOR\n(C5)"),
        ("F6", "FECHA OPERACIÓN\n(C6)"),
        ("G6", "GLOSA DE OPERACIÓN\n(C7)"),
        ("H6", "MONTO TOTAL\nFLUJO INGRESO O EGRESO\n(C8)"),
        ("I6", "MONTO QUE AFECTA\nLA BASE IMPONIBLE\n(C9)"),
    ]

    for celda, texto in headers:
        cell = ws[celda]
        cell.value = texto
        cell.font = Font(name="Arial", bold=True, color=BLANCO, size=8)
        cell.fill = PatternFill("solid", start_color=AZUL_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borde_fino

    ws.row_dimensions[6].height = 45

    # ── Ancho de columnas ──────────────────────────────────────────────────
    anchos = {"A": 8, "B": 10, "C": 14, "D": 14, "E": 14, "F": 12,
               "G": 40, "H": 16, "I": 16}
    for col, ancho in anchos.items():
        ws.column_dimensions[col].width = ancho

    # ── Datos del libro ────────────────────────────────────────────────────
    fila_inicio = 7
    for idx, row in df.iterrows():
        fila = fila_inicio + idx
        tipo_op = row["Tipo Operación"]
        fondo_fila = GRIS_CLARO if idx % 2 == 0 else BLANCO

        if tipo_op == 0:
            fondo_fila = AZUL_CLARO

        # Fecha
        fecha = row["Fecha Operación"]
        fecha_str = fecha.strftime("%d/%m/%Y") if pd.notna(fecha) else ""

        # Tipo doc — mostrar nombre si existe
        tipo_doc = row["Tipo Documento"]
        try:
            tipo_doc_int = int(float(str(tipo_doc))) if str(tipo_doc).strip() else ""
            tipo_doc_str = str(tipo_doc_int) if tipo_doc_int else ""
        except Exception:
            tipo_doc_str = str(tipo_doc) if tipo_doc else ""

        valores = [
            row["N° Correlativo"],
            tipo_op,
            str(row["N° Documento"]),
            tipo_doc_str,
            str(row["RUT Emisor"]) if row["RUT Emisor"] else "",
            fecha_str,
            str(row["Glosa de Operación"]),
            row["C8"] if row["C8"] != 0 else None,
            row["C9"] if row["C9"] != 0 else None,
        ]

        alineaciones = ["center", "center", "center", "center", "center",
                        "center", "left", "right", "right"]
        formatos = [None, None, None, None, None, None, None,
                    '#,##0', '#,##0']

        for col_idx, (val, aln, fmt) in enumerate(zip(valores, alineaciones, formatos), 1):
            cell = ws.cell(row=fila, column=col_idx, value=val)
            cell.font = Font(name="Arial", size=9,
                             bold=(tipo_op == 0))
            cell.fill = PatternFill("solid", start_color=fondo_fila)
            cell.alignment = Alignment(horizontal=aln, vertical="center", wrap_text=True)
            cell.border = borde_fino
            if fmt:
                cell.number_format = fmt

        ws.row_dimensions[fila].height = 16

    # ── Fila separadora ────────────────────────────────────────────────────
    fila_sep = fila_inicio + len(df)
    ws.merge_cells(f"A{fila_sep}:I{fila_sep}")
    ws[f"A{fila_sep}"].fill = PatternFill("solid", start_color=AZUL_HEADER)
    ws.row_dimensions[fila_sep].height = 6

    # ── SALDOS Y TOTALES LIBRO DE CAJA (formato oficial SII) ──────────────
    NEGRO = "000000"
    AMARILLO_HEADER = "FFC000"
    AMARILLO_CLARO = "FFF2CC"

    # Fila 1: Título principal
    f1 = fila_sep + 1
    ws.merge_cells(f"A{f1}:I{f1}")
    cell_t = ws[f"A{f1}"]
    cell_t.value = "SALDOS Y TOTALES LIBRO DE CAJA"
    cell_t.font = Font(name="Arial", bold=True, color=NEGRO, size=10)
    cell_t.fill = PatternFill("solid", start_color=AMARILLO_HEADER)
    cell_t.alignment = Alignment(horizontal="center", vertical="center")
    cell_t.border = borde_fino
    for col in range(2, 10):
        ws.cell(row=f1, column=col).border = borde_fino
        ws.cell(row=f1, column=col).fill = PatternFill("solid", start_color=AMARILLO_HEADER)
    ws.row_dimensions[f1].height = 22

    # Fila 2: Sub-encabezados de grupo
    f2 = f1 + 1
    ws.merge_cells(f"A{f2}:F{f2}")
    cell_flujo = ws[f"A{f2}"]
    cell_flujo.value = "FLUJO DE INGRESOS Y EGRESOS"
    cell_flujo.font = Font(name="Arial", bold=True, color=NEGRO, size=9)
    cell_flujo.fill = PatternFill("solid", start_color=AMARILLO_CLARO)
    cell_flujo.alignment = Alignment(horizontal="center", vertical="center")
    cell_flujo.border = borde_fino
    for col in range(2, 7):
        ws.cell(row=f2, column=col).border = borde_fino
        ws.cell(row=f2, column=col).fill = PatternFill("solid", start_color=AMARILLO_CLARO)

    ws.merge_cells(f"G{f2}:I{f2}")
    cell_bi = ws[f"G{f2}"]
    cell_bi.value = "MONTOS QUE AFECTAN LA BASE IMPONIBLE"
    cell_bi.font = Font(name="Arial", bold=True, color=NEGRO, size=9)
    cell_bi.fill = PatternFill("solid", start_color=AMARILLO_CLARO)
    cell_bi.alignment = Alignment(horizontal="center", vertical="center")
    cell_bi.border = borde_fino
    for col in range(8, 10):
        ws.cell(row=f2, column=col).border = borde_fino
        ws.cell(row=f2, column=col).fill = PatternFill("solid", start_color=AMARILLO_CLARO)
    ws.row_dimensions[f2].height = 20

    # Fila 3: Etiquetas de cada total
    f3 = f2 + 1
    labels_row3 = [
        (1, 2, "TOTAL MONTO\nFLUJO DE\nINGRESOS"),
        (3, 4, "TOTAL MONTO FLUJO\nDE EGRESOS"),
        (5, 6, "SALDO FLUJO DE\nCAJA"),
        (7, 7, "INGRESOS"),
        (8, 8, "EGRESOS"),
        (9, 9, "RESULTADO\nNETO"),
    ]
    for col_ini, col_fin, texto in labels_row3:
        if col_ini != col_fin:
            ws.merge_cells(start_row=f3, start_column=col_ini, end_row=f3, end_column=col_fin)
        cell = ws.cell(row=f3, column=col_ini)
        cell.value = texto
        cell.font = Font(name="Arial", bold=True, color=NEGRO, size=8)
        cell.fill = PatternFill("solid", start_color=AMARILLO_CLARO)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borde_fino
        if col_ini != col_fin:
            for c in range(col_ini + 1, col_fin + 1):
                ws.cell(row=f3, column=c).border = borde_fino
                ws.cell(row=f3, column=c).fill = PatternFill("solid", start_color=AMARILLO_CLARO)
    ws.row_dimensions[f3].height = 38

    # Fila 4: Códigos C10-C15
    f4 = f3 + 1
    codigos = [
        (1, 2, "C10"),
        (3, 4, "C11"),
        (5, 6, "C12"),
        (7, 7, "C13"),
        (8, 8, "C14"),
        (9, 9, "C15"),
    ]
    for col_ini, col_fin, codigo in codigos:
        if col_ini != col_fin:
            ws.merge_cells(start_row=f4, start_column=col_ini, end_row=f4, end_column=col_fin)
        cell = ws.cell(row=f4, column=col_ini)
        cell.value = codigo
        cell.font = Font(name="Arial", bold=True, color=NEGRO, size=9)
        cell.fill = PatternFill("solid", start_color=AMARILLO_CLARO)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = borde_fino
        if col_ini != col_fin:
            for c in range(col_ini + 1, col_fin + 1):
                ws.cell(row=f4, column=c).border = borde_fino
                ws.cell(row=f4, column=c).fill = PatternFill("solid", start_color=AMARILLO_CLARO)
    ws.row_dimensions[f4].height = 18

    # Fila 5: Valores C10-C15
    f5 = f4 + 1
    valores_totales = [
        (1, 2, totales["total_ingresos"]),
        (3, 4, totales["total_egresos"]),
        (5, 6, totales["saldo_flujo"]),
        (7, 7, totales["ing_bi"]),
        (8, 8, totales["egr_bi"]),
        (9, 9, totales["resultado_neto"]),
    ]
    for col_ini, col_fin, valor in valores_totales:
        if col_ini != col_fin:
            ws.merge_cells(start_row=f5, start_column=col_ini, end_row=f5, end_column=col_fin)
        cell = ws.cell(row=f5, column=col_ini)
        cell.value = valor
        cell.font = Font(name="Arial", bold=True, color=NEGRO, size=10)
        cell.fill = PatternFill("solid", start_color=BLANCO)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = borde_fino
        cell.number_format = '#,##0'
        if col_ini != col_fin:
            for c in range(col_ini + 1, col_fin + 1):
                ws.cell(row=f5, column=c).border = borde_fino
                ws.cell(row=f5, column=c).fill = PatternFill("solid", start_color=BLANCO)
    ws.row_dimensions[f5].height = 22

    # ── Pie de página ──────────────────────────────────────────────────────
    fila_pie = f5 + 2
    ws.merge_cells(f"A{fila_pie}:I{fila_pie}")
    ws[f"A{fila_pie}"].value = (
        "Documento generado conforme al Art. 14 D N°3 y N°8(a) LIR — Régimen Pro Pyme — SII Chile"
    )
    ws[f"A{fila_pie}"].font = Font(name="Arial", italic=True, size=8, color="666666")
    ws[f"A{fila_pie}"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[fila_pie].height = 14

    # ── Guardar en memoria ─────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# EXPORTACIÓN PDF (solo totales)
# ─────────────────────────────────────────────────────────────────────────────
def exportar_pdf_totales(
    totales: dict,
    rut_empresa: str,
    nombre_empresa: str,
    periodo: str,
) -> bytes:
    """Genera un PDF con la tabla de Saldos y Totales del Libro de Caja."""

    def fmt(v):
        signo = "-" if v < 0 else ""
        return f"{signo}${abs(v):,.0f}".replace(",", ".")

    def safe(text: str) -> str:
        """Reemplaza caracteres fuera de latin-1 para compatibilidad con Helvetica."""
        reemplazos = {
            "\u2014": "-", "\u2013": "-",  # em-dash, en-dash
            "\u00b0": "o.",                 # °  → o.
            "\u201c": '"', "\u201d": '"',   # comillas tipográficas
            "\u2018": "'", "\u2019": "'",
        }
        for viejo, nuevo in reemplazos.items():
            text = text.replace(viejo, nuevo)
        # Asegurar que todo sea encodeable a latin-1
        return text.encode("latin-1", errors="replace").decode("latin-1")

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.add_page()
    pdf.set_auto_page_break(auto=False)

    # ── Colores ────────────────────────────────────────────────────────────
    DORADO = (255, 192, 0)
    AMARILLO = (255, 242, 204)
    AZUL_OSC = (31, 56, 100)
    BLANCO = (255, 255, 255)
    NEGRO = (0, 0, 0)

    page_w = 297  # A4 landscape
    margen = 15
    tabla_w = page_w - 2 * margen

    # ── Título SII ─────────────────────────────────────────────────────────
    pdf.set_fill_color(*AZUL_OSC)
    pdf.set_text_color(*BLANCO)
    pdf.set_font("Helvetica", "B", 11)
    pdf.set_xy(margen, 12)
    pdf.cell(tabla_w, 12,
             safe("ANEXO 3. LIBRO DE CAJA - REGIMEN PRO PYME (Art. 14 D No.3 y No.8 LIR)"),
             border=1, align="C", fill=True)

    # ── Datos empresa ──────────────────────────────────────────────────────
    pdf.set_text_color(*NEGRO)
    pdf.set_fill_color(242, 242, 242)
    pdf.set_font("Helvetica", "B", 9)

    y_datos = 28
    col_label_w = 50
    col_val_w = tabla_w - col_label_w

    for label, valor in [("PERIODO:", periodo),
                         ("RUT:", rut_empresa),
                         ("NOMBRE / RAZON SOCIAL:", nombre_empresa)]:
        pdf.set_xy(margen, y_datos)
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(col_label_w, 7, safe(label), border=1, fill=True)
        pdf.set_font("Helvetica", "", 9)
        pdf.cell(col_val_w, 7, safe(valor), border=1, fill=True)
        y_datos += 7

    # ── Tabla de totales ───────────────────────────────────────────────────
    y_tabla = y_datos + 8

    # Anchos de las 6 columnas
    col_w = tabla_w / 6

    # Fila 1: Título principal
    pdf.set_xy(margen, y_tabla)
    pdf.set_fill_color(*DORADO)
    pdf.set_text_color(*NEGRO)
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(tabla_w, 10, "SALDOS Y TOTALES LIBRO DE CAJA",
             border=1, align="C", fill=True)
    y_tabla += 10

    # Fila 2: Sub-grupos
    pdf.set_xy(margen, y_tabla)
    pdf.set_fill_color(*AMARILLO)
    pdf.set_font("Helvetica", "B", 9)
    pdf.cell(col_w * 3, 8, "FLUJO DE INGRESOS Y EGRESOS",
             border=1, align="C", fill=True)
    pdf.cell(col_w * 3, 8, "MONTOS QUE AFECTAN LA BASE IMPONIBLE",
             border=1, align="C", fill=True)
    y_tabla += 8

    # Fila 3: Etiquetas
    pdf.set_xy(margen, y_tabla)
    pdf.set_fill_color(*AMARILLO)
    pdf.set_font("Helvetica", "B", 7)
    labels = [
        "TOTAL MONTO\nFLUJO DE\nINGRESOS",
        "TOTAL MONTO\nFLUJO DE\nEGRESOS",
        "SALDO FLUJO\nDE CAJA",
        "INGRESOS",
        "EGRESOS",
        "RESULTADO\nNETO",
    ]
    h_labels = 14
    for lb in labels:
        x_pos = pdf.get_x()
        pdf.set_fill_color(*AMARILLO)
        pdf.rect(x_pos, y_tabla, col_w, h_labels, "DF")
        # Centrar texto multi-linea
        lines = lb.split("\n")
        line_h = h_labels / max(len(lines), 1)
        for j, line in enumerate(lines):
            pdf.set_xy(x_pos, y_tabla + j * line_h + (h_labels - len(lines) * line_h) / 2)
            pdf.cell(col_w, line_h, line, align="C")
        pdf.set_xy(x_pos + col_w, y_tabla)
    y_tabla += h_labels

    # Fila 4: Códigos C10-C15
    pdf.set_xy(margen, y_tabla)
    pdf.set_fill_color(*AMARILLO)
    pdf.set_font("Helvetica", "B", 9)
    for code in ["C10", "C11", "C12", "C13", "C14", "C15"]:
        pdf.cell(col_w, 7, code, border=1, align="C", fill=True)
    y_tabla += 7

    # Fila 5: Valores
    pdf.set_xy(margen, y_tabla)
    pdf.set_fill_color(*BLANCO)
    pdf.set_font("Helvetica", "B", 11)
    valores = [
        totales["total_ingresos"],
        totales["total_egresos"],
        totales["saldo_flujo"],
        totales["ing_bi"],
        totales["egr_bi"],
        totales["resultado_neto"],
    ]
    for val in valores:
        pdf.cell(col_w, 10, fmt(val), border=1, align="C", fill=True)
    y_tabla += 10

    # ── Pie de página ──────────────────────────────────────────────────────
    pdf.set_xy(margen, y_tabla + 6)
    pdf.set_font("Helvetica", "I", 7)
    pdf.set_text_color(120, 120, 120)
    pdf.cell(tabla_w, 5,
             safe("Documento generado conforme al Art. 14 D No.3 y No.8(a) LIR - Regimen Pro Pyme - SII Chile"),
             align="C")

    buffer = io.BytesIO()
    pdf.output(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# VALIDACIONES
# ─────────────────────────────────────────────────────────────────────────────
def validar_libro(df: pd.DataFrame) -> list[str]:
    advertencias = []

    # Duplicados
    df_ops = df[df["Tipo Operación"] != 0]
    duplicados = df_ops[df_ops.duplicated(
        subset=["N° Documento", "Tipo Documento", "Tipo Operación"], keep=False
    )]
    if not duplicados.empty:
        folios = duplicados["N° Documento"].unique()[:5]
        advertencias.append(
            f"⚠️ Posibles documentos duplicados detectados: {', '.join(str(f) for f in folios)}"
        )

    # Base imponible > total
    mask_problema = df["C9"] > df["C8"] + 1
    if mask_problema.any():
        advertencias.append(
            "⚠️ Existen registros donde la Base Imponible supera el Monto Total. Verifique los datos."
        )

    # Correlativo sin saltos
    correlativos = df["N° Correlativo"].tolist()
    esperado = list(range(1, len(correlativos) + 1))
    if correlativos != esperado:
        advertencias.append("⚠️ El correlativo tiene saltos o irregularidades.")

    return advertencias


# ─────────────────────────────────────────────────────────────────────────────
# INTERFAZ STREAMLIT
# ─────────────────────────────────────────────────────────────────────────────
def main():
    st.set_page_config(
        page_title="Libro de Caja — Pro Pyme SII",
        page_icon="📒",
        layout="wide",
        initial_sidebar_state="expanded",
    )

    # ── CSS personalizado ──────────────────────────────────────────────────
    st.markdown("""
    <style>
        .main { background-color: #f8f9fa; }
        .stButton>button {
            background-color: #1F3864;
            color: white;
            border-radius: 6px;
            font-weight: bold;
        }
        .stButton>button:hover { background-color: #2E5090; }
        .metric-card {
            background: white;
            border-radius: 8px;
            padding: 16px;
            box-shadow: 0 1px 4px rgba(0,0,0,0.1);
            text-align: center;
        }
        .header-sii {
            background: linear-gradient(90deg, #1F3864 0%, #2E5090 100%);
            color: white;
            padding: 20px 30px;
            border-radius: 8px;
            margin-bottom: 20px;
        }
        div[data-testid="stDataFrame"] {font-size: 12px;}
        .warning-box {
            background-color: #fff3cd;
            border-left: 4px solid #ffc107;
            padding: 10px;
            margin: 5px 0;
            border-radius: 4px;
        }
    </style>
    """, unsafe_allow_html=True)

    # ── Header ─────────────────────────────────────────────────────────────
    st.markdown("""
    <div class="header-sii">
        <h2 style="margin:0;">📒 LIBRO DE CAJA — RÉGIMEN PRO PYME</h2>
        <p style="margin:4px 0 0 0; opacity:0.85; font-size:0.9em;">
            Art. 14 Letra D) N°3 y N°8(a) Ley sobre Impuesto a la Renta — SII Chile
        </p>
    </div>
    """, unsafe_allow_html=True)

    # ── Sidebar — Parámetros empresa ───────────────────────────────────────
    with st.sidebar:
        logo_path = os.path.join(os.path.dirname(__file__), "logo.png")
        if os.path.exists(logo_path):
            st.image(logo_path, width=120, use_container_width=False)
        else:
            st.warning("Logo no encontrado")
        st.markdown("---")
        st.subheader("🏢 Datos de la Empresa")

        rut_empresa = st.text_input("RUT Empresa", placeholder="76.123.456-7")
        nombre_empresa = st.text_input("Nombre / Razón Social",
                                       placeholder="EMPRESA EJEMPLO LTDA.")
        periodo = st.text_input("Año Comercial",
                                value=str(datetime.now().year),
                                placeholder="2025",
                                help="Ingrese el año en formato YYYY",
                                max_chars=4)

        st.markdown("---")
        st.subheader("💰 Saldo Inicial")
        saldo_inicial = st.number_input(
            "Saldo Inicial ($)",
            min_value=0,
            value=0,
            step=1000,
            format="%d",
            help="Ingrese el saldo inicial del libro de caja para el período",
            key="saldo_inicial_input",
        )

        st.markdown("---")
        st.subheader("ℹ️ Tipo Operación")
        st.markdown("""
        | Código | Descripción |
        |--------|-------------|
        | **0** | Saldo Inicial |
        | **1** | Flujo Ingreso |
        | **2** | Flujo Egreso |
        """)

    # ── Carga de archivos ──────────────────────────────────────────────────
    st.markdown("## 📂 Carga de Archivos")

    col1, col2, col3 = st.columns(3)

    with col1:
        st.markdown("### 🟢 Ventas (Facturas)")
        st.caption("Archivos CSV del registro de ventas (facturas electrónicas)")
        archivos_ventas = st.file_uploader(
            "Cargar CSV de Ventas",
            type=["csv"],
            accept_multiple_files=True,
            key="ventas",
            help="Suba uno o más archivos CSV con el detalle de facturas de venta"
        )

    with col2:
        st.markdown("### 🟡 Resumen Ventas (Boletas)")
        st.caption("CSV con resumen de boletas y comprobantes electrónicos")
        archivos_resumen = st.file_uploader(
            "Cargar CSV Resumen Ventas",
            type=["csv"],
            accept_multiple_files=True,
            key="resumen",
            help="Suba el archivo de resumen de ventas con boletas"
        )

    with col3:
        st.markdown("### 🔴 Compras")
        st.caption("Archivos CSV del registro de compras (un archivo por mes)")
        archivos_compras = st.file_uploader(
            "Cargar CSV de Compras",
            type=["csv"],
            accept_multiple_files=True,
            key="compras",
            help="Suba uno o más archivos CSV con el detalle de compras"
        )

    # ── Datos pegados (F29 y Honorarios) dentro de expanders ──────────────
    with st.expander("🔵 Pagos F29 (Pegar Datos)", expanded=False):
        st.markdown("""
        <div style="background:#e8f0fe; border-left:4px solid #1a73e8; padding:10px; border-radius:4px; margin-bottom:10px; font-size:0.85em;">
            <strong>📋 Formato esperado</strong> (6 campos por fila, separados por comas):<br>
            <code>C2, Folio, Descripción, Fecha, Detalle, Monto</code><br>
            <em>Ejemplo:</em> <code>2,8091536376,Formulario F-29,06/02/2025,Pago del F-29,"$151,077"</code>
        </div>
        """, unsafe_allow_html=True)
        texto_f29 = st.text_area(
            "Pegar datos F29",
            height=150,
            key="f29_texto",
            placeholder="2,8091536376,Formulario F-29,06/02/2025,Pago del F-29,\"$151,077\"\n2,8116188706,Formulario F-29,12/03/2025,Pago del F-29,\"$139,050\"\n...",
            help="Pegue aquí los datos de los pagos de F29."
        )

    with st.expander("🟣 Honorarios (Pegar Datos)", expanded=False):
        st.markdown("""
        <div style="background:#f3e8fe; border-left:4px solid #9b59b6; padding:10px; border-radius:4px; margin-bottom:10px; font-size:0.85em;">
            <strong>📋 Formato esperado</strong> (8 campos por fila, separados por comas):<br>
            <code>C2, N° Documento, Tipo Doc, RUT, Fecha, Nombre, Pagado, Bruto</code><br>
            <em>Ejemplo:</em> <code>2,1389,Boleta de Honorarios Electrónica,13064992-0,05/09/2024,ALEJANDRO PATRICIO DUQUE SOTO,108684,108684</code>
        </div>
        """, unsafe_allow_html=True)
        texto_honorarios = st.text_area(
            "Pegar datos Honorarios",
            height=150,
            key="honorarios_texto",
            placeholder="2,1389,Boleta de Honorarios Electrónica,13064992-0,05/09/2024,ALEJANDRO PATRICIO DUQUE SOTO,108684,108684\n...",
            help="Pegue aquí los datos de las boletas de honorarios."
        )

    # ── Botón procesar ─────────────────────────────────────────────────────
    st.markdown("---")
    col_btn1, col_btn2, col_btn3 = st.columns([1, 2, 1])
    with col_btn2:
        boton_generar = st.button("⚙️ GENERAR LIBRO DE CAJA", use_container_width=True)

    # Cuando se presiona el botón, regenerar desde cero y guardar en session_state
    if boton_generar:
        tiene_texto_f29 = texto_f29 and texto_f29.strip()
        tiene_texto_honorarios = texto_honorarios and texto_honorarios.strip()
        if not archivos_ventas and not archivos_resumen and not archivos_compras and not tiene_texto_f29 and not tiene_texto_honorarios:
            st.error("❌ Debe cargar al menos un archivo o pegar datos para procesar.")
        else:
            if not rut_empresa or not nombre_empresa or not periodo:
                st.warning("⚠️ Complete los datos de la empresa en el panel lateral antes de procesar.")

            with st.spinner("Procesando archivos..."):
                df_ventas = pd.DataFrame()
                df_compras = pd.DataFrame()

                if archivos_ventas or archivos_resumen:
                    try:
                        df_ventas = procesamiento_ventas(
                            archivos_ventas or [],
                            archivos_resumen or [],
                            periodo or "",
                        )
                    except Exception as e:
                        st.error(f"❌ Error procesando ventas: {e}")

                if archivos_compras:
                    try:
                        df_compras = procesamiento_compras(archivos_compras)
                    except Exception as e:
                        st.error(f"❌ Error procesando compras: {e}")

                if tiene_texto_honorarios:
                    try:
                        df_honorarios = procesar_texto_honorarios(texto_honorarios)
                        if not df_honorarios.empty:
                            if df_compras.empty:
                                df_compras = df_honorarios
                            else:
                                df_compras = pd.concat([df_compras, df_honorarios], ignore_index=True)
                    except Exception as e:
                        st.error(f"❌ Error procesando honorarios: {e}")

                if tiene_texto_f29:
                    try:
                        df_f29 = procesar_texto_f29(texto_f29)
                        if not df_f29.empty:
                            if df_compras.empty:
                                df_compras = df_f29
                            else:
                                df_compras = pd.concat([df_compras, df_f29], ignore_index=True)
                    except Exception as e:
                        st.error(f"❌ Error procesando datos F29: {e}")

                df_libro_nuevo = generar_libro_caja(
                    df_ventas, df_compras,
                    float(saldo_inicial),
                    rut_empresa or "00.000.000-0",
                    nombre_empresa or "SIN NOMBRE",
                    periodo or "SIN PERÍODO",
                )

                advertencias = validar_libro(df_libro_nuevo)

                # Guardar en session_state (fuente de verdad persistente)
                st.session_state["df_libro"] = df_libro_nuevo
                st.session_state["totales"] = calcular_totales(df_libro_nuevo)
                st.session_state["advertencias"] = advertencias
                st.session_state["editor_version"] = 0  # Reiniciar el editor
                st.session_state["rut_empresa"] = rut_empresa
                st.session_state["nombre_empresa"] = nombre_empresa
                st.session_state["periodo"] = periodo

    # ── Mostrar tabla si ya fue generada ──────────────────────────────────
    if "df_libro" not in st.session_state:
        # ── Estado inicial: instrucciones ──────────────────────────────────
        st.info("""
        **Cómo usar esta aplicación:**

        1. **Complete los datos de la empresa** en el panel lateral (RUT, Nombre, Período)
        2. **Ingrese el saldo inicial de caja** (único dato manual requerido)
        3. **Cargue los archivos CSV** del SII:
           - 📁 *Ventas*: Archivo de facturas electrónicas emitidas
           - 📁 *Resumen Ventas*: Archivo con resumen de boletas y comprobantes
           - 📁 *Compras*: Archivo de facturas de compra recibidas
        4. **Haga clic en "Generar Libro de Caja"**
        5. **Descargue el Excel** con el formato oficial SII
        """)

        col1, col2 = st.columns(2)
        with col1:
            st.markdown("""
            **📌 Columnas del Libro de Caja (SII):**
            | Columna | Descripción |
            |---------|-------------|
            | C1 | N° Correlativo |
            | C2 | Tipo Operación (0/1/2) |
            | C3 | N° Documento |
            | C4 | Tipo Documento |
            | C5 | RUT Emisor |
            | C6 | Fecha Operación |
            | C7 | Glosa |
            | C8 | Monto Total Flujo |
            | C9 | Monto Base Imponible |
            """)
        with col2:
            st.markdown("""
            **📌 Tipos de Operación:**
            | Código | Significado |
            |--------|-------------|
            | 0 | Saldo Inicial |
            | 1 | Ingreso de Caja |
            | 2 | Egreso de Caja |

            **📌 Régimen Pro Pyme:**
            - El IVA **no** afecta la base imponible
            - Base imponible = Neto + Exento
            - Régimen Art. 14 D N°3 y N°8(a) LIR
            """)
        return

    # ── Leer desde session_state ───────────────────────────────────────────
    df_libro = st.session_state["df_libro"]
    totales   = st.session_state["totales"]
    advertencias = st.session_state.get("advertencias", [])
    # Usar siempre los valores actuales del sidebar (pueden haber cambiado sin regenerar)
    rut_empresa_ss    = rut_empresa or st.session_state.get("rut_empresa", "00.000.000-0")
    nombre_empresa_ss = nombre_empresa or st.session_state.get("nombre_empresa", "SIN NOMBRE")
    periodo_ss        = periodo or st.session_state.get("periodo", "SIN PERÍODO")
    editor_ver        = st.session_state.get("editor_version", 0)

    # ── Advertencias ───────────────────────────────────────────────────────
    if advertencias:
        st.markdown("### ⚠️ Advertencias de Validación")
        for adv in advertencias:
            st.warning(adv)

    st.markdown("## 📋 Libro de Caja")

    # Preparar df para mostrar (no modifica session_state)
    df_display = df_libro.copy()
    df_display["Fecha Operación"] = df_display["Fecha Operación"].apply(
        lambda x: x.date() if pd.notna(x) else None
    )

    def tipo_doc_label(v):
        try:
            codigo = int(float(str(v))) if str(v).strip() else 0
            nombre = TIPO_DOC_NOMBRES.get(codigo, "")
            return f"({codigo}) {nombre}" if nombre else str(v)
        except Exception:
            return str(v)

    df_display["Tipo Documento"] = df_display["Tipo Documento"].apply(tipo_doc_label)
    df_display["C9"] = df_display["C9"].apply(
        lambda x: f"$ {x:,.0f}".replace(",", ".") if x else ""
    )

    df_show = df_display.rename(columns={
        "N° Correlativo": "C1 N° Correlativo",
        "Tipo Operación": "C2 Tipo Operación",
        "N° Documento": "C3 N° Documento",
        "Tipo Documento": "C4 Tipo Documento",
        "RUT Emisor": "C5 RUT Emisor",
        "Fecha Operación": "C6 Fecha Operación",
        "Glosa de Operación": "C7 Glosa",
        "C8": "C8 Monto Total",
        "C9": "C9 Base Imponible",
    }).drop(columns=["_origen"], errors="ignore")

    def color_fila(row):
        if row["C2 Tipo Operación"] == 0:
            return ["background-color: #BDD7EE"] * len(row)
        elif row["C2 Tipo Operación"] == 1:
            return ["background-color: #E2EFDA"] * len(row)
        elif row["C2 Tipo Operación"] == 2:
            return ["background-color: #FCE4D6"] * len(row)
        return [""] * len(row)

    st.markdown("### ✏️ Editar Libro de Caja")
    st.caption("Edita el saldo inicial (C8) y las fechas (C6). Al confirmar cambios, la tabla se reordenará y el correlativo se recalculará.")

    # Clave dinámica: cuando se incrementa editor_version, el editor se reinicia
    # con los datos ya ordenados de session_state
    edited_df = st.data_editor(
        df_show.style.apply(color_fila, axis=1),
        use_container_width=True,
        height=520,
        disabled=[
            "C1 N° Correlativo", "C2 Tipo Operación", "C3 N° Documento",
            "C4 Tipo Documento", "C5 RUT Emisor",
            "C7 Glosa", "C9 Base Imponible"
        ],
        column_config={
            "C6 Fecha Operación": st.column_config.DateColumn(
                "C6 Fecha Operación",
                help="Haz clic para cambiar la fecha. Luego presiona 'Aplicar cambios'.",
                format="DD/MM/YYYY",
            ),
            "C8 Monto Total": st.column_config.NumberColumn(
                "C8 Monto Total",
                help="Ingrese monto total",
                min_value=0,
                step=1,
                format="$ %d",
            ),
        },
        key=f"editor_tabla_{editor_ver}"
    )

    # Botón para aplicar cambios explícitamente (reordena y recalcula)
    col_ap1, col_ap2, col_ap3 = st.columns([1, 2, 1])
    with col_ap2:
        if st.button("🔄 Aplicar cambios y reordenar", use_container_width=True, key="btn_aplicar"):
            if edited_df is not None and not edited_df.empty:
                try:
                    df_trabajo = df_libro.copy()

                    # 1) Actualizar saldo inicial
                    fila_saldo = edited_df[edited_df["C2 Tipo Operación"] == 0]
                    if not fila_saldo.empty:
                        nuevo_saldo = float(fila_saldo.iloc[0]["C8 Monto Total"] or 0)
                        idx_saldo = df_trabajo[df_trabajo["Tipo Operación"] == 0].index
                        if not idx_saldo.empty:
                            df_trabajo.loc[idx_saldo[0], "C8"] = nuevo_saldo

                    # 2) Propagar fechas editadas
                    for i, row in edited_df.iterrows():
                        nueva_fecha = row.get("C6 Fecha Operación")
                        if nueva_fecha is not None and i < len(df_trabajo):
                            try:
                                df_trabajo.at[i, "Fecha Operación"] = pd.Timestamp(nueva_fecha)
                            except Exception:
                                pass

                    # 3) Reordenar: saldo inicial siempre primero, resto por fecha
                    saldo_rows = df_trabajo[df_trabajo["Tipo Operación"] == 0].copy()
                    otros_rows = df_trabajo[df_trabajo["Tipo Operación"] != 0].sort_values(
                        "Fecha Operación", na_position="first"
                    )
                    df_trabajo = pd.concat([saldo_rows, otros_rows], ignore_index=True)

                    # 4) Recalcular correlativo
                    df_trabajo["N° Correlativo"] = range(1, len(df_trabajo) + 1)

                    # 5) Guardar en session_state y forzar reinicio del editor
                    st.session_state["df_libro"] = df_trabajo
                    st.session_state["totales"] = calcular_totales(df_trabajo)
                    st.session_state["editor_version"] = editor_ver + 1
                    st.rerun()

                except Exception as e:
                    st.error(f"Error al aplicar cambios: {e}")

    # Leyenda
    col_l1, col_l2, col_l3 = st.columns(3)
    col_l1.markdown("🔵 **Saldo Inicial**")
    col_l2.markdown("🟢 **Flujo Ingreso**")
    col_l3.markdown("🔴 **Flujo Egreso**")

    # ── Totales — formato oficial SII ──────────────────────────────────────
    st.markdown("## 📊 Saldos y Totales Libro de Caja")

    def fmt_clp(v):
        signo = "-" if v < 0 else ""
        return f"{signo}${abs(v):,.0f}".replace(",", ".")

    # Tabla HTML con el formato del SII
    saldo_color = "#27ae60" if totales["saldo_flujo"] >= 0 else "#e74c3c"
    resultado_color = "#27ae60" if totales["resultado_neto"] >= 0 else "#e74c3c"

    st.markdown(f"""
    <div style="overflow-x:auto;">
    <table style="width:100%; border-collapse:collapse; font-family:Arial, sans-serif; font-size:14px;">
        <thead>
            <tr>
                <th colspan="6" style="background:#FFC000; color:#000; text-align:center; padding:10px; border:1px solid #ccc; font-size:15px;">
                    SALDOS Y TOTALES LIBRO DE CAJA
                </th>
            </tr>
            <tr>
                <th colspan="3" style="background:#FFF2CC; color:#000; text-align:center; padding:8px; border:1px solid #ccc;">
                    FLUJO DE INGRESOS Y EGRESOS
                </th>
                <th colspan="3" style="background:#FFF2CC; color:#000; text-align:center; padding:8px; border:1px solid #ccc;">
                    MONTOS QUE AFECTAN LA BASE IMPONIBLE
                </th>
            </tr>
            <tr>
                <th style="background:#FFF2CC; color:#000; text-align:center; padding:6px; border:1px solid #ccc; font-size:12px;">
                    TOTAL MONTO<br>FLUJO DE<br>INGRESOS
                </th>
                <th style="background:#FFF2CC; color:#000; text-align:center; padding:6px; border:1px solid #ccc; font-size:12px;">
                    TOTAL MONTO<br>FLUJO DE<br>EGRESOS
                </th>
                <th style="background:#FFF2CC; color:#000; text-align:center; padding:6px; border:1px solid #ccc; font-size:12px;">
                    SALDO FLUJO<br>DE CAJA
                </th>
                <th style="background:#FFF2CC; color:#000; text-align:center; padding:6px; border:1px solid #ccc; font-size:12px;">
                    INGRESOS
                </th>
                <th style="background:#FFF2CC; color:#000; text-align:center; padding:6px; border:1px solid #ccc; font-size:12px;">
                    EGRESOS
                </th>
                <th style="background:#FFF2CC; color:#000; text-align:center; padding:6px; border:1px solid #ccc; font-size:12px;">
                    RESULTADO<br>NETO
                </th>
            </tr>
            <tr>
                <th style="background:#FFF2CC; color:#000; text-align:center; padding:4px; border:1px solid #ccc; font-size:11px;">C10</th>
                <th style="background:#FFF2CC; color:#000; text-align:center; padding:4px; border:1px solid #ccc; font-size:11px;">C11</th>
                <th style="background:#FFF2CC; color:#000; text-align:center; padding:4px; border:1px solid #ccc; font-size:11px;">C12</th>
                <th style="background:#FFF2CC; color:#000; text-align:center; padding:4px; border:1px solid #ccc; font-size:11px;">C13</th>
                <th style="background:#FFF2CC; color:#000; text-align:center; padding:4px; border:1px solid #ccc; font-size:11px;">C14</th>
                <th style="background:#FFF2CC; color:#000; text-align:center; padding:4px; border:1px solid #ccc; font-size:11px;">C15</th>
            </tr>
        </thead>
        <tbody>
            <tr>
                <td style="text-align:center; padding:12px; border:1px solid #ccc; font-weight:bold; font-size:15px; color:#27ae60;">
                    {fmt_clp(totales["total_ingresos"])}
                </td>
                <td style="text-align:center; padding:12px; border:1px solid #ccc; font-weight:bold; font-size:15px; color:#e74c3c;">
                    {fmt_clp(totales["total_egresos"])}
                </td>
                <td style="text-align:center; padding:12px; border:1px solid #ccc; font-weight:bold; font-size:15px; color:{saldo_color};">
                    {fmt_clp(totales["saldo_flujo"])}
                </td>
                <td style="text-align:center; padding:12px; border:1px solid #ccc; font-weight:bold; font-size:15px; color:#27ae60;">
                    {fmt_clp(totales["ing_bi"])}
                </td>
                <td style="text-align:center; padding:12px; border:1px solid #ccc; font-weight:bold; font-size:15px; color:#e74c3c;">
                    {fmt_clp(totales["egr_bi"])}
                </td>
                <td style="text-align:center; padding:12px; border:1px solid #ccc; font-weight:bold; font-size:15px; color:{resultado_color};">
                    {fmt_clp(totales["resultado_neto"])}
                </td>
            </tr>
        </tbody>
    </table>
    </div>
    """, unsafe_allow_html=True)

    # ── Descarga ────────────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("## 💾 Exportar")

    excel_bytes = exportar_excel(
        df_libro, totales,
        rut_empresa_ss,
        nombre_empresa_ss,
        periodo_ss,
    )
    nombre_base = f"LibroCaja_{rut_empresa_ss.replace('.', '').replace('-', '')}_{periodo_ss}"

    col_dl1, col_dl2 = st.columns(2)
    with col_dl1:
        st.download_button(
            label="📥 Descargar Libro de Caja (Excel)",
            data=excel_bytes,
            file_name=f"{nombre_base}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with col_dl2:
        pdf_bytes = exportar_pdf_totales(
            totales,
            rut_empresa_ss,
            nombre_empresa_ss,
            periodo_ss,
        )
        st.download_button(
            label="📄 Descargar Totales (PDF)",
            data=pdf_bytes,
            file_name=f"{nombre_base}_Totales.pdf",
            mime="application/pdf",
            use_container_width=True,
        )

    st.success(f"✅ Libro de Caja con {len(df_libro)} registros. Usa '🔄 Aplicar cambios' para confirmar ediciones.")


if __name__ == "__main__":
    main()

